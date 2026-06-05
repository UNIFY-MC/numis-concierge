# -*- coding: utf-8 -*-
# Cria uma cópia .xlsx do XLS dos escudos com as moedas DUVIDOSAS (match fraco ou
# sem-match) a VERMELHO, para o dono corrigir à mão. Lê os duvidosos de
# scripts/.escudos-uncertos.json (folha+linha de origem).
import xlrd, json
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font

SRC = r'assets/coleçaodadospai/Colecção de Moedas de 1917 a 2002.XLS'
DST = r'assets/coleçaodadospai/Colecção escudos - REVISÃO (vermelho=corrigir).xlsx'

uncertos = json.load(open(r'scripts/.escudos-uncertos.json', encoding='utf-8'))
porFolha = {}
for u in uncertos:
    porFolha.setdefault(u['sheet'], set()).add(u['row'])

RED = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
REDFONT = Font(color='9C0006', bold=True)

wb_in = xlrd.open_workbook(SRC)
wb = Workbook()
wb.remove(wb.active)
total_marcadas = 0
for sn in wb_in.sheet_names():
    ws_in = wb_in.sheet_by_name(sn)
    ws = wb.create_sheet(title=sn[:31])
    marcar = porFolha.get(sn, set())
    for r in range(ws_in.nrows):
        vermelho = r in marcar
        if vermelho:
            total_marcadas += 1
        for c in range(ws_in.ncols):
            v = ws_in.cell_value(r, c)
            cell = ws.cell(row=r + 1, column=c + 1, value=v if v != '' else None)
            if vermelho:
                cell.fill = RED
                cell.font = REDFONT

wb.save(DST)
print(f'✓ {total_marcadas} linhas a vermelho · {DST}')
