# -*- coding: utf-8 -*-
# Marca a VERMELHO no ficheiro dos euros (.xlsx) tudo o que é DÚBIO ou NÃO IMPORTADO:
#  - circulação possuída (qtd>=1) que NÃO está na nossa coleção (cobertura em
#    scripts/.euro-coverage.json: pais|facial|ano);
#  - a zona das COMEMORATIVAS (o dono confere à mão; não foram importadas daqui).
# Gera uma cópia: "...REVISÃO (vermelho).xlsx".
import json, re, unicodedata
from openpyxl import load_workbook
from openpyxl.styles import PatternFill, Font

SRC = r'assets/coleçaodadospai/Coleccao_Euros_Paises novo2.xlsx'
DST = r'assets/coleçaodadospai/Coleccao_Euros - REVISÃO (vermelho=dubio_ou_nao_importado).xlsx'
RED = PatternFill(start_color='FFC7CE', end_color='FFC7CE', fill_type='solid')
REDF = Font(color='9C0006', bold=True)

cov = set(json.load(open(r'scripts/.euro-coverage.json', encoding='utf-8')))

def na(s):  # normaliza (sem acentos, minúsculas)
    return ''.join(c for c in unicodedata.normalize('NFD', str(s)) if unicodedata.category(c) != 'Mn').lower().strip()

PAIS = {
    'alemanha': 'de', 'andorra': 'ad', 'austria': 'at', 'belgica': 'be', 'bulgaria': 'bg',
    'chipre': 'cy', 'croacia': 'hr', 'eslovaquia': 'sk', 'eslovenia': 'si', 'espanha': 'es',
    'estonia': 'ee', 'finlandia': 'fi', 'franca': 'fr', 'grecia': 'gr', 'holanda': 'nl',
    'irlanda': 'ie', 'italia': 'it', 'letonia': 'lv', 'lituania': 'lt', 'luxemburgo': 'lu',
    'malta': 'mt', 'portugal': 'pt', 'san marino': 'sm', 'vaticano': 'va',
}
def paisDe(sheet):
    n = na(sheet)
    for k, v in PAIS.items():
        if n.startswith(k):
            return v
    return None

def facialDe(label):
    n = na(label)
    m = re.match(r'(\d+)\s*(centimo|cent|euro)', n)
    if not m:
        return None
    num = int(m.group(1))
    return num / 100 if 'cent' in m.group(2) else float(num)

def formatoDe(sheet):
    n = na(sheet)
    if 'bebe' in n or 'bebé' in n:
        return 'carteira_bebe'
    if 'carteira' in n:
        return 'carteira_fdc'
    return 'bnc'

wb = load_workbook(SRC)
tot_circ = 0; tot_com = 0
faltam = []
for ws in wb.worksheets:
    if 'ndice' in na(ws.title):
        continue
    pais = paisDe(ws.title)
    # 1) localizar linha de anos ("Ano & data") e linha "Comemorativa" / denom
    linha_ano = None
    for r in range(1, min(ws.max_row, 30) + 1):
        a = na(ws.cell(r, 1).value)
        if a.startswith('ano') and ('data' in a or 'ano &' in a):
            linha_ano = r; break
    # 2) marcar COMEMORATIVAS (zona entre o topo e a linha de anos: rótulos "Moed." / "Comemorativa")
    fim_com = (linha_ano or 16) - 1
    for r in range(1, fim_com + 1):
        a = na(ws.cell(r, 1).value)
        if 'comemorativ' in a or a.startswith('moed.') or a.startswith('moed '):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).fill = RED; ws.cell(r, c).font = REDF
            tot_com += 1
    if not linha_ano or not pais:
        continue
    # 3) anos por par de colunas (qtdCol par-esquerda, rotuloCol = qtdCol+1)
    anos = {}
    for qcol in range(2, ws.max_column + 1, 2):
        y = ws.cell(linha_ano, qcol + 1).value
        try:
            yi = int(float(y))
            if 1999 <= yi <= 2035:
                anos[qcol] = yi
        except Exception:
            pass
    # 4) linhas de denominação (a seguir à linha de anos)
    for r in range(linha_ano + 1, min(ws.max_row, linha_ano + 12) + 1):
        fac = facialDe(ws.cell(r, 1).value)
        if fac is None:
            continue
        for qcol, ano in anos.items():
            try:
                q = int(float(ws.cell(r, qcol).value))
            except Exception:
                q = 0
            if q < 1:
                continue
            # chave igual à do JS: inteiros sem ".0" (1€/2€ → "1"/"2", cêntimos → "0.05")
            fkey = str(int(fac)) if fac == int(fac) else str(fac)
            chave = f'{pais}|{fkey}|{ano}'
            if chave not in cov:  # possuído no ficheiro mas não na coleção → vermelho
                ws.cell(r, qcol).fill = RED; ws.cell(r, qcol).font = REDF
                ws.cell(r, qcol + 1).fill = RED; ws.cell(r, qcol + 1).font = REDF
                tot_circ += 1
                faltam.append({'pais': pais, 'facial': fac, 'ano': ano, 'qua': q, 'formato': formatoDe(ws.title)})

wb.save(DST)
json.dump(faltam, open(r'scripts/.euro-faltam.json', 'w', encoding='utf-8'), ensure_ascii=False)
print(f'  {len(faltam)} a importar → scripts/.euro-faltam.json')
print(f'✓ circulação não-importada a vermelho: {tot_circ} células · comemorativas (linhas) marcadas: {tot_com}')
print(f'  → {DST}')
